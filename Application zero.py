import openpyxl
# kindly change the path of the file as it is in your computer.
# be sure you install openpyxl library
countries = {"egypt": "egypt.xlsx",
             "canada": "canda.xlsx",
             "china": "china.xlsx",
             "usa": "usa.xlsx"
              }
total_pop = []
country_exist = 0
# To help the user to use the application.

print("please enter the number of the process you want.")
print("1- Enter a new country to load its file ")
print("2-Display the population of each state and total population of the country")
print("3-Display the state with the highest population and lowest population")
print("4- exit")

while True:
    # input number of task the user need
    print("________________________")
    number = int(input("enter the number: "))
    if (number > 4) or (number < 1):
        number = int(input("enter correct number: "))
    # task number 1
    if number == 1:
        # add name and path of the file the user need to add
        name = input("please enter the name of the country you want to add.")
        path = input("please enter the path you want to add -be sure it is a correct one - ")
        countries[name] = path
        number = int(input("enter the number: "))
        if (number > 4) or (number < 2):
            number = int(input("enter correct number: "))
        if number == 2:
            country = str(input("enter the needed country"))
            for i in countries.keys():
                if i == country:
                    country_exist = 1
                    path = countries[i]
                    wb_obj = openpyxl.load_workbook(path)
                    sheet_obj = wb_obj.active
                    m_row = sheet_obj.max_row
                    max_col = sheet_obj.max_column
                    # print the population of each state
                    for i in range(1, m_row + 1):
                        cell_city_obj = sheet_obj.cell(row=i, column=1)
                        cell_pop_obj = sheet_obj.cell(row=i, column=2)
                        print(cell_city_obj.value, ": ", cell_pop_obj.value)
                        # print total population of the country
                        total_pop.append(cell_pop_obj.value)
                    print("total population = ", sum(total_pop))
            if country_exist == 0:
                print("country is not available")
            else:
                country_exist = 0


        elif number == 3:
            country = str(input("enter the needed country"))
            for i in countries.keys():
                if i == country:
                    country_exist = 1
                    path = countries[i]
                    wb_obj = openpyxl.load_workbook(path)
                    sheet_obj = wb_obj.active
                    max_col = sheet_obj.max_column
                    # print the most population city
                    for i in range(1, max_col + 1):
                        cell_obj = sheet_obj.cell(row=1, column=2)
                        cell1_obj = sheet_obj.cell(row=1, column=1)
                    print("The most population city is: ", cell1_obj.value, ":", cell_obj.value)

                    # print the least population city
                    last_row = sheet_obj.max_row
                    last_col = sheet_obj.max_column
                    last_col1_a_value = sheet_obj.cell(column=1, row=last_row).value
                    last_col_a_value = sheet_obj.cell(column=2, row=last_row).value
                    print("The least population city is: ", last_col1_a_value, ":", last_col_a_value)
            if country_exist == 0:
                print("country is not available")
            else:
                country_exist = 0

        else:
            print("end program")
            exit()

    # task 2
    elif number == 2:
        country = str(input("enter the needed country"))
        for i in countries.keys():
            if i == country:
                country_exist = 1
                path = countries[i]
                wb_obj = openpyxl.load_workbook(path)
                sheet_obj = wb_obj.active
                m_row = sheet_obj.max_row
                max_col = sheet_obj.max_column
                # print the population of each state
                for i in range(1, m_row + 1):
                    cell_city_obj = sheet_obj.cell(row=i, column=1)
                    cell_pop_obj = sheet_obj.cell(row=i, column=2)
                    print(cell_city_obj.value, ": ", cell_pop_obj.value)
                    # print total population of the country
                    total_pop.append(cell_pop_obj.value)
                # print total population of the country
                print("total population = ", sum(total_pop))
        if country_exist == 0:
            print("country is not available")
        else:
            country_exist = 0


    elif number == 3:
        country = str(input("enter the needed country"))
        for i in countries.keys():
            if i == country:
                country_exist = 1
                path = countries[i]
                wb_obj = openpyxl.load_workbook(path)
                sheet_obj = wb_obj.active
                max_col = sheet_obj.max_column
                # print the most population city
                for i in range(1, max_col + 1):
                    cell_obj = sheet_obj.cell(row=1, column=2)
                    cell1_obj = sheet_obj.cell(row=1, column=1)
                print("The most population city is: ", cell1_obj.value, ":", cell_obj.value)

                # print the least population city
                last_row = sheet_obj.max_row
                last_col = sheet_obj.max_column
                last_col1_a_value = sheet_obj.cell(column=1, row=last_row).value
                last_col_a_value = sheet_obj.cell(column=2, row=last_row).value
                print("The least population city is: ", last_col1_a_value, ":", last_col_a_value)
        if country_exist == 0:
            print("country is not available")
        else:
            country_exist = 0

    else:
        print("end program")
        exit()
